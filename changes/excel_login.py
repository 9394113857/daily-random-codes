import openpyxl
import pyautogui
import csv

# Function to display sheet data
def display_sheet_data(sheet, rows_to_display):
    # Get headers
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    # Print headers
    print("|", end="")
    for header in headers:
        print(f" {header} |", end="")
    print("\n" + "-" * (len(headers) * 10))

    # Print selected rows
    for row_number in rows_to_display:
        print("|", end="")
        row_data = [cell.value for cell in sheet[row_number] if cell.value is not None]  # Filter out None values
        for item in row_data:
            print(f" {item} |", end="")
        print()

    # Add sleep time after displaying sheet data
    pyautogui.sleep(5)

    # Call login function for the last row's credentials
    last_row = rows_to_display[-1]
    login(sheet.cell(row=last_row, column=2).value, sheet.cell(row=last_row, column=3).value)

# Function to perform login actions
def login(username, password):
    # Type the username and press Enter
    pyautogui.write(username)
    pyautogui.press('enter')
    print(f"2. Typing Username: {username}")
    pyautogui.sleep(5)

    # Type the password and press Enter
    pyautogui.write(password)
    pyautogui.press('enter')
    print("3. Typing Password")
    pyautogui.sleep(10)

    # Press 'f6' key again (optional, remove if not needed)
    pyautogui.write(['f6'])

    # Type the YouTube URL and press Enter
    YouTube = "www.youtube.com"
    pyautogui.write(YouTube)
    pyautogui.press('enter')
    print("4. Entering into YouTube")

    # Add a delay or sleep to allow time for the page to load (adjust sleep time as needed)
    pyautogui.sleep(10)

# Function to perform logout actions
def logout():
    pyautogui.write(['f6'])
    logout_url = "https://accounts.google.com/logout"
    pyautogui.write(logout_url)
    pyautogui.press('enter')
    print("5. Logged out successfully")

# Sleep for 5 seconds before starting to allow time to focus on the target window
pyautogui.sleep(5)

# Hardcode the Excel file name
xlsx_file_name = "cc.xlsx"

# Read available sheets
try:
    wb = openpyxl.load_workbook(xlsx_file_name)
    sheet_names = wb.sheetnames
    print("Available sheets:")
    for i, sheet_name in enumerate(sheet_names, start=1):
        print(f"{i}. {sheet_name}")

    # Ask user for sheet choice
    sheet_choice = input("Enter the sheet number to load: ")
    if sheet_choice.isdigit():
        sheet_choice = int(sheet_choice) - 1
        if 0 <= sheet_choice < len(sheet_names):
            selected_sheet = wb[sheet_names[sheet_choice]]

            # Skip the first row (assuming it contains headers)
            rows_to_display = range(2, selected_sheet.max_row + 1)

            # Ask user if they want to specify rows
            rows_to_display_option = input("Do you want to specify rows to display? (yes/no): ")
            if rows_to_display_option.lower() == "yes":
                rows_input = input("Enter row numbers separated by commas (e.g., 1,2,3): ")
                if rows_input:
                    rows_to_display = [int(row.strip()) + 1 for row in rows_input.split(",")]

            # Display selected rows
            display_sheet_data(selected_sheet, rows_to_display)

            # Call logout function after displaying sheet data
            logout()
        else:
            print("Invalid sheet number.")
    else:
        print("Invalid input. Please enter a valid sheet number.")
except Exception as e:
    print(f"Error: {e}")
