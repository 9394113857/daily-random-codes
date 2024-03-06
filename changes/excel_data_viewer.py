import openpyxl

def display_sheet_data(sheet, rows_to_display):
    # Get headers
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    # Print headers
    print("|", end="")
    for header in headers:
        print(f" {header} |", end="")
    print("\n" + "-" * (len(headers) * 10))

    # Print selected rows
    for row_number in rows_to_display:
        print("|", end="")
        row_data = [cell.value for cell in sheet[row_number] if cell.value is not None]  # Filter out None values
        for item in row_data:
            print(f" {item} |", end="")
        print()

# Hardcode the Excel file name
xlsx_file_name = "cc.xlsx"

# Read available sheets
try:
    wb = openpyxl.load_workbook(xlsx_file_name)
    sheet_names = wb.sheetnames
    print("Available sheets:")
    for i, sheet_name in enumerate(sheet_names, start=1):
        print(f"{i}. {sheet_name}")

    # Ask user for sheet choice
    sheet_choice = input("Enter the sheet number to load: ")
    if sheet_choice.isdigit():
        sheet_choice = int(sheet_choice) - 1
        if 0 <= sheet_choice < len(sheet_names):
            selected_sheet = wb[sheet_names[sheet_choice]]

            # Ask user if they want to specify rows
            rows_to_display = input("Do you want to specify rows to display? (yes/no): ")
            if rows_to_display.lower() == "yes":
                rows_input = input("Enter row numbers separated by commas (e.g., 1,2,3): ")
                if rows_input:
                    rows_to_display = [int(row.strip()) for row in rows_input.split(",")]
                    rows_to_display = [row + 1 for row in rows_to_display]  # Increment each row number by 1
                    # Display selected rows
                    display_sheet_data(selected_sheet, rows_to_display)
                else:
                    print("No rows specified. Displaying all rows:")
                    display_sheet_data(selected_sheet, range(2, selected_sheet.max_row + 1))  # Display all rows except headers
            else:
                # Display all rows
                print("Displaying all rows:")
                display_sheet_data(selected_sheet, range(2, selected_sheet.max_row + 1))  # Display all rows except headers
        else:
            print("Invalid sheet number.")
    else:
        print("Invalid input. Please enter a valid sheet number.")
except Exception as e:
    print(f"Error: {e}")
