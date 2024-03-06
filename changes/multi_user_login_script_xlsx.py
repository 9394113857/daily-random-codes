import openpyxl
import pyautogui

# Function to perform login actions
def login(username, password):
    # Type the username and press Enter
    pyautogui.write(username)
    pyautogui.press('enter')
    print(f"2. Typing Username: {username}")
    pyautogui.sleep(5)

    # Type the password and press Enter
    pyautogui.write(password)
    pyautogui.press('enter')
    print("3. Typing Password")
    pyautogui.sleep(10)

    # Press 'f6' key again (optional, remove if not needed)
    pyautogui.write(['f6'])

    # Type the YouTube URL and press Enter
    YouTube = "www.youtube.com"
    pyautogui.write(YouTube)
    pyautogui.press('enter')
    print("4. Entering into YouTube")

    # Add a delay or sleep to allow time for the page to load (adjust sleep time as needed)
    pyautogui.sleep(10)

# Function to perform logout actions
def logout():
    pyautogui.write(['f6'])
    logout_url = "https://accounts.google.com/logout"
    pyautogui.write(logout_url)
    pyautogui.press('enter')
    print("5. Logged out successfully")

# Sleep for 5 seconds before starting to allow time to focus on the target window
pyautogui.sleep(5)

# Press 'f6' key (optional, remove if not needed)
pyautogui.write(['f6'])

# Type the URL directly using pyautogui
link = "https://accounts.google.com/signin"
pyautogui.write(link)
pyautogui.press('enter')  # Press Enter to navigate to the URL
print("1. Typing URL")

# Wait for the login page to load (adjust sleep time as needed)
pyautogui.sleep(10)

# Hardcode the Excel file name
xlsx_file_name = "cc.xlsx"

# Read sno, username, and password from Excel file
try:
    wb = openpyxl.load_workbook(xlsx_file_name)
    sheet_names = wb.sheetnames
    print("Available sheets:")
    for sheet_name in sheet_names:
        print(sheet_name)

    # Prompt the user to enter the sheet name manually
    sheet_name = input("Enter the sheet name: ")

    ws = wb[sheet_name]
    rows_to_process = input("Do you want to specify rows to process? (yes/no): ")
    if rows_to_process.lower() == "yes":
        row_numbers = input("Enter the row numbers separated by commas (e.g., 1,2,3 etc): ")
        rows = [int(row.strip()) for row in row_numbers.split(",")]
    else:
        rows = None

    # Add sleep time after taking row numbers input
    pyautogui.sleep(5)

    # Loop through each row in the Excel sheet
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        sno, username, password = row

        # Check if the row needs to be processed based on user input
        if rows is None or i in rows:
            print(f"\nExecuting tasks for Row {i} (SNO: {sno})")
            login(username, password)  # Perform login actions
            logout()  # Perform logout actions

            # Wait for some time before moving to the next row (adjust sleep time as needed)
            pyautogui.sleep(5)

    print("\nAll tasks completed for all users.")
except Exception as e:
    print(f"Error: {e}")
