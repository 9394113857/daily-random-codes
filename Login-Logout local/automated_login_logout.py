import openpyxl
import pyautogui

# Function to perform login actions
def login(username, password):
    # Type the username and press Enter
    pyautogui.write(username)
    pyautogui.press('enter')
    print(f"2. Typing Username: {username}")
    pyautogui.sleep(5)

    # Type the password and press Enter
    pyautogui.write(password)
    pyautogui.press('enter')
    print("3. Typing Password")
    pyautogui.sleep(5)

    # Press 'f6' key again (optional, remove if not needed)
    pyautogui.press('f6')

    # Add additional sleep time before accessing YouTube
    pyautogui.sleep(10)  # Additional sleep time for stability

    # Type the YouTube URL and press Enter
    YouTube = "www.youtube.com"
    pyautogui.write(YouTube)
    pyautogui.press('enter')
    print("4. Entering into YouTube")

    # Add a delay or sleep to allow time for the page to load
    pyautogui.sleep(10)  # Increased sleep time for YouTube page load

# Function to perform logout actions
def logout():
    pyautogui.press('f6')
    logout_url = "https://accounts.google.com/logout"
    pyautogui.write(logout_url)
    pyautogui.press('enter')
    print("5. Logged out successfully")
    pyautogui.sleep(5)  # Sleep time after logout

# Sleep for 5 seconds before starting to allow time to focus on the target window
pyautogui.sleep(5)

# Press 'f6' key (optional, remove if not needed)
pyautogui.press('f6')

# Type the URL directly using pyautogui
link = "https://accounts.google.com/signin"
pyautogui.write(link)
pyautogui.press('enter')  # Press Enter to navigate to the URL
print("1. Typing URL")

# Wait for the login page to load (adjust sleep time as needed)
pyautogui.sleep(5)

# Corrected Excel file path
xlsx_file_name = "D:\\daily-random-codes-main\\Login-Logout local\\cc.xlsx"

# Read sno, username, and password from Excel file
try:
    wb = openpyxl.load_workbook(xlsx_file_name)
    sheet_names = wb.sheetnames
    print("Available sheets:", sheet_names)

    # Prompt the user to enter the sheet number manually
    sheet_number = int(input("Enter the sheet number: "))  # Convert input to integer

    # Select the sheet based on the entered sheet number
    ws = wb.worksheets[sheet_number - 1]  # Subtract 1 to adjust for zero-based indexing

    rows_to_process = input("Do you want to specify rows to process? (yes/no): ")
    if rows_to_process.lower() == "yes":
        row_numbers = input("Enter the row numbers separated by commas (e.g., 1,6): ")
        rows = [int(row.strip()) + 1 for row in row_numbers.split(",")]  # Adjusting row numbers as per user input
    else:
        rows = None

    # Add sleep time after taking row numbers input
    pyautogui.sleep(5)

    # Loop through each row in the Excel sheet
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        sno, username, password = row

        # Check if the row needs to be processed based on user input
        if rows is None or i in rows:
            print(f"\nExecuting tasks for Row {i} (SNO: {sno})")
            login(username, password)  # Perform login actions
            logout()  # Perform logout actions

            # Wait for some time before moving to the next row
            pyautogui.sleep(5)

            # Check if there are more users in the loop
            if i < len(list(ws.iter_rows(values_only=True))):
                # Press 'f6' key (optional, remove if not needed)
                pyautogui.press('f6')

                # Type the URL directly using pyautogui
                link = "https://accounts.google.com/signin"
                pyautogui.write(link)
                pyautogui.press('enter')  # Press Enter to navigate to the URL
                print("1. Typing URL")

                # Wait for the login page to load (adjust sleep time as needed)
                pyautogui.sleep(5)

    print("\nAll tasks completed for all users.")
except Exception as e:
    print(f"Error: {e}")
